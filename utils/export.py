"""
Export utility: Excel/CSV export and DB backup.
"""
import os
import shutil
import csv
from datetime import datetime
from typing import List, Dict


def export_to_csv(data: List[Dict], filepath: str):
    if not data:
        raise ValueError("No data to export.")
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)


def export_to_excel(data: List[Dict], filepath: str, sheet_name: str = "Students"):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            wb.save(filepath)
            return

        headers = list(data[0].keys())

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Data rows
        alt_fill = PatternFill("solid", fgColor="F0F0F0")
        for row_idx, row in enumerate(data, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if fill:
                    cell.fill = fill

        # Auto-fit column widths
        for col_idx, header in enumerate(headers, 1):
            max_len = len(header) + 2
            for row in data:
                val = str(row.get(header, "") or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        ws.row_dimensions[1].height = 22
        wb.save(filepath)

    except ImportError:
        # Fallback to CSV
        csv_path = filepath.replace(".xlsx", ".csv")
        export_to_csv(data, csv_path)
        raise ImportError(f"openpyxl not installed. Exported as CSV to: {csv_path}")


def backup_database(db_path: str, backup_dir: str) -> str:
    os.makedirs(backup_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = os.path.join(backup_dir, f"library_backup_{timestamp}.db")
    shutil.copy2(db_path, backup_path)
    return backup_path
